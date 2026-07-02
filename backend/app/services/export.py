"""测评分析导出：把选中的 evaluation 打包成 zip。

包含：
- README.md         说明文件
- summary.xlsx      汇总表：
    · 总体对比 sheet —— 行=任务，列分「准确率/耗时/Token均值/CoT均值」四组，每组各模型一列
    · 明细 sheet     —— 每条 evaluation 一行（含 token 分位统计）
- charts.html       静态对比图（Chart.js CDN）
- raw/<eval_id>/    每个 evaluation 的 details_path 目录全量复制

token 长度统计：用轻量 `tokenizers` 库直接加载 tokenizer.json，对 infer 产物
predictions/*.jsonl 里的 prediction 文本做编码（full=整段、cot=<think> 推理段）。
tokenizer 缺失或 jsonl 找不到时，token 列留空，不影响其余导出。
"""
import io
import json
import re
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Iterator

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from backend.app.config import get_settings
from backend.app.models import (
    Batch, Evaluation, Job, Model, Prediction, Task,
)


# ── token 统计 ────────────────────────────────────────────────────────

def _load_tokenizer():
    """加载 qwen tokenizer（tokenizer.json）。缺库或缺文件时返回 None。"""
    try:
        from tokenizers import Tokenizer
    except ImportError:
        return None
    path = get_settings().workspace_dir / "tokenizers" / "qwen3" / "tokenizer.json"
    if not path.exists():
        return None
    try:
        return Tokenizer.from_file(str(path))
    except Exception:
        return None


def _extract_cot(text: str) -> str:
    """提取 <think>…</think> 推理段；被截断时尽量兜底。"""
    if not text:
        return ""
    tl = text.lower()
    o, c = tl.find("<think>"), tl.find("</think>")
    if o != -1 and c != -1 and o < c:
        return text[o + 7:c]
    if c != -1:
        return text[:c]
    if o != -1:
        return text[o + 7:]
    return ""


def _pred_text(obj: dict):
    """从 prediction jsonl 的一行里取出模型输出文本。"""
    p = obj.get("prediction")
    if isinstance(p, str):
        return p
    if isinstance(p, dict):
        return p.get("prediction")
    return None


def _percentile(sorted_vals: list[int], q: float):
    if not sorted_vals:
        return None
    k = (len(sorted_vals) - 1) * q
    lo = int(k)
    hi = min(lo + 1, len(sorted_vals) - 1)
    return round(sorted_vals[lo] + (sorted_vals[hi] - sorted_vals[lo]) * (k - lo), 1)


def _token_stats(output_task_id: str | None, tokenizer) -> dict | None:
    """对一个 infer 产物（output_task_id）下的 predictions jsonl 编码、聚合 token 统计。"""
    if not tokenizer or not output_task_id:
        return None
    base = get_settings().workspace_dir / "outputs" / output_task_id / "details"
    files = sorted(base.glob("**/predictions/**/*.jsonl"))
    full_texts, cot_texts = [], []
    for jf in files:
        try:
            with open(jf, encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        obj = json.loads(line)
                    except json.JSONDecodeError:
                        continue
                    txt = _pred_text(obj)
                    if not txt:
                        continue
                    full_texts.append(txt)
                    cot = _extract_cot(txt)
                    if cot:
                        cot_texts.append(cot)
        except OSError:
            continue
    if not full_texts:
        return None

    def lens(texts):
        if not texts:
            return []
        return sorted(len(e.ids) for e in tokenizer.encode_batch(texts))

    full = lens(full_texts)
    cot = lens(cot_texts)

    def mean(xs):
        return round(sum(xs) / len(xs), 1) if xs else None

    return {
        "samples": len(full),
        "full_mean": mean(full),
        "full_p90": _percentile(full, 0.90),
        "full_max": full[-1] if full else None,
        "cot_mean": mean(cot),
        "cot_p90": _percentile(cot, 0.90),
        "cot_max": cot[-1] if cot else None,
    }


# ── 主流程 ────────────────────────────────────────────────────────────

def build_analysis_zip(db: Session, evaluation_ids: list[int], title: str,
                       view_id: int | None = None) -> Iterator[bytes]:
    """构建对比 zip 并以流式返回（生成器，逐块产出，适合 StreamingResponse）。"""
    tokenizer = _load_tokenizer()
    rows = _collect_rows(db, evaluation_ids, tokenizer)

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("README.md", _readme(title, view_id, rows))
        zf.writestr("summary.xlsx", _build_xlsx(rows))
        zf.writestr("charts.html", _build_charts_html(title, rows))
        _add_raw_details(zf, rows, tokenizer)
    buf.seek(0)
    while True:
        chunk = buf.read(64 * 1024)
        if not chunk:
            break
        yield chunk


def _collect_rows(db: Session, evaluation_ids: list[int], tokenizer=None) -> list[dict]:
    ids = evaluation_ids or []
    if not ids:
        return []
    rows: list[dict] = []
    for eid in ids:
        ev = db.get(Evaluation, eid)
        if not ev:
            continue
        pred = db.get(Prediction, ev.prediction_id) if ev.prediction_id else None
        model = db.get(Model, pred.model_id) if pred else None
        task = db.get(Task, pred.task_id) if pred else None
        job = db.get(Job, ev.job_id) if ev.job_id else None
        batch = db.get(Batch, job.batch_id) if job and job.batch_id else None
        rows.append({
            "evaluation_id": ev.id,
            "version_label": ev.version_label,
            "status": ev.status,
            "accuracy": ev.accuracy,
            "num_samples": ev.num_samples,
            "duration_sec": ev.duration_sec,
            "eval_version": ev.eval_version,
            "model_id": pred.model_id if pred else None,
            "model_name": model.name if model else None,
            "task_id": pred.task_id if pred else None,
            "task_key": task.key if task else None,
            "batch_id": batch.id if batch else None,
            "batch_name": batch.name if batch else None,
            "details_path": ev.details_path,
            "output_task_id": pred.output_task_id if pred else None,
            "finished_at": ev.finished_at.isoformat() if ev.finished_at else None,
            "token": _token_stats(pred.output_task_id if pred else None, tokenizer),
        })
    return rows


def _uniq(seq):
    """保序去重，元素为 (id, label) 元组。"""
    seen, out = set(), []
    for k, v in seq:
        if k not in seen:
            seen.add(k)
            out.append((k, v))
    return out


def _build_xlsx(rows: list[dict]) -> bytes:
    wb = Workbook()
    _build_overview_sheet(wb.active, rows)
    _build_detail_sheet(wb.create_sheet("明细"), rows)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_overview_sheet(ws, rows: list[dict]):
    """总体对比：行=任务，列分四组（准确率/耗时/Token均值/CoT均值），每组各模型一列。"""
    ws.title = "总体对比"
    models = _uniq((r["model_id"], r["model_name"] or str(r["model_id"])) for r in rows)
    tasks = _uniq((r["task_id"], r["task_key"] or str(r["task_id"])) for r in rows)
    idx = {(r["model_id"], r["task_id"]): r for r in rows}

    groups = [
        ("准确率(%)", lambda r: r["accuracy"]),
        ("耗时(s)", lambda r: round(r["duration_sec"], 1) if r["duration_sec"] is not None else None),
        ("Token均值", lambda r: r["token"]["full_mean"] if r.get("token") else None),
        ("CoT均值", lambda r: r["token"]["cot_mean"] if r.get("token") else None),
    ]

    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    hdr_fill = PatternFill("solid", fgColor="EEF2F7")

    # 第 1 行：分组标题（每组横跨 len(models) 列，合并）；第 2 行：任务 + 各组下模型名
    ws.cell(1, 1).fill = hdr_fill
    c0 = ws.cell(2, 1, "任务"); c0.font = bold; c0.fill = hdr_fill
    col = 2
    for gname, _ in groups:
        start = col
        gc = ws.cell(1, col, gname); gc.font = bold; gc.alignment = center; gc.fill = hdr_fill
        for _, mname in models:
            mc = ws.cell(2, col, mname); mc.font = bold; mc.fill = hdr_fill
            col += 1
        if col - 1 >= start:
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=col - 1)

    # 数据行
    r_i = 3
    for tid, tkey in tasks:
        ws.cell(r_i, 1, tkey).font = bold
        col = 2
        for _, getter in groups:
            for mid, _ in models:
                row = idx.get((mid, tid))
                ws.cell(r_i, col, getter(row) if row else None)
                col += 1
        r_i += 1

    ws.column_dimensions["A"].width = 22


def _build_detail_sheet(ws, rows: list[dict]):
    headers = [
        "模型", "任务", "批次", "版本", "状态", "准确率(%)", "样本数", "耗时(s)",
        "Token均值", "Token_P90", "Token_Max", "CoT均值", "CoT_P90", "CoT_Max",
    ]
    bold = Font(bold=True)
    for j, h in enumerate(headers, start=1):
        ws.cell(1, j, h).font = bold
    for i, r in enumerate(rows, start=2):
        tk = r.get("token") or {}
        vals = [
            r["model_name"] or r["model_id"],
            r["task_key"] or r["task_id"],
            r["batch_name"] or r["batch_id"],
            r["version_label"] or "",
            r["status"],
            r["accuracy"],
            r["num_samples"],
            round(r["duration_sec"], 1) if r["duration_sec"] is not None else None,
            tk.get("full_mean"), tk.get("full_p90"), tk.get("full_max"),
            tk.get("cot_mean"), tk.get("cot_p90"), tk.get("cot_max"),
        ]
        for j, v in enumerate(vals, start=1):
            ws.cell(i, j, v)


def _build_charts_html(title: str, rows: list[dict]) -> str:
    """生成单页 HTML，用 Chart.js CDN 渲染对比柱图 + 表格。"""
    labels = [
        f"{r['model_name'] or r['model_id']} · {r['task_key'] or r['task_id']} · {r['version_label'] or ''}"
        for r in rows
    ]
    accuracies = [r["accuracy"] if r["accuracy"] is not None else 0 for r in rows]
    durations = [round(r["duration_sec"], 2) if r["duration_sec"] is not None else 0 for r in rows]
    tokens = [(r["token"]["full_mean"] if r.get("token") and r["token"].get("full_mean") is not None else 0) for r in rows]

    data_json = json.dumps({
        "labels": labels, "accuracies": accuracies,
        "durations": durations, "tokens": tokens,
    }, ensure_ascii=False)

    table_rows_html = "".join(
        f"<tr><td>{r['model_name'] or ''}</td><td>{r['task_key'] or ''}</td>"
        f"<td>{r['batch_name'] or ''}</td><td>{r['version_label'] or ''}</td>"
        f"<td>{r['status']}</td>"
        f"<td>{r['accuracy'] if r['accuracy'] is not None else '—'}</td>"
        f"<td>{r['num_samples'] if r['num_samples'] is not None else '—'}</td>"
        f"<td>{round(r['duration_sec'], 2) if r['duration_sec'] is not None else '—'}</td>"
        f"<td>{(r['token'] or {}).get('full_mean', '—') if r.get('token') else '—'}</td></tr>"
        for r in rows
    )

    now = datetime.utcnow().isoformat()
    title = title or "对比报告"
    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
  <meta charset="UTF-8" />
  <title>{title} · 对比报告</title>
  <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
  <style>
    body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
            padding: 32px; max-width: 1200px; margin: 0 auto; color: #222; }}
    h1 {{ margin-bottom: 4px; }}
    .meta {{ color: #888; font-size: 13px; margin-bottom: 24px; }}
    .chart-box {{ margin: 24px 0; padding: 16px; border: 1px solid #eee; border-radius: 8px; }}
    table {{ width: 100%; border-collapse: collapse; margin-top: 24px; font-size: 13px; }}
    th, td {{ border: 1px solid #e5e7eb; padding: 6px 10px; text-align: left; }}
    th {{ background: #f9fafb; }}
  </style>
</head>
<body>
  <h1>{title}</h1>
  <div class="meta">生成时间：{now} · 共 {len(rows)} 条 evaluation</div>

  <div class="chart-box"><h3>准确率对比 (%)</h3><canvas id="acc"></canvas></div>
  <div class="chart-box"><h3>耗时对比 (s)</h3><canvas id="dur"></canvas></div>
  <div class="chart-box"><h3>输出 Token 均值</h3><canvas id="tok"></canvas></div>

  <h3>明细表</h3>
  <table>
    <thead><tr>
      <th>模型</th><th>任务</th><th>批次</th><th>版本</th>
      <th>状态</th><th>准确率</th><th>样本数</th><th>耗时(s)</th><th>Token均值</th>
    </tr></thead>
    <tbody>{table_rows_html}</tbody>
  </table>

  <script>
    const data = {data_json};
    const mk = (id, label, arr, color, max) => new Chart(document.getElementById(id), {{
      type: 'bar',
      data: {{ labels: data.labels, datasets: [{{ label, data: arr, backgroundColor: color }}] }},
      options: {{ scales: {{ y: {{ beginAtZero: true, ...(max ? {{ max }} : {{}}) }} }} }}
    }});
    mk('acc', '准确率 (%)', data.accuracies, 'rgba(59,130,246,0.6)', 100);
    mk('dur', '耗时 (s)', data.durations, 'rgba(245,158,11,0.6)');
    mk('tok', 'Token 均值', data.tokens, 'rgba(16,185,129,0.6)');
  </script>
</body>
</html>"""


def _fmt_value(v):
    """格式化单元格值：dict/list 转 JSON、None→null、剔除 Excel 非法控制字符。"""
    if v is None:
        return "null"
    if isinstance(v, (dict, list)):
        s = json.dumps(v, ensure_ascii=False, indent=2)
    else:
        s = str(v)
    return re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", s)


def _details_jsonl_to_xlsx(jf: Path, tokenizer) -> bytes | None:
    """把一个 *_details.jsonl 转成 xlsx（列同 aggregate_eval_reports.py）。

    列：eval_res / eval_details / origin_prompt / prediction / gold / full_tokens / cot_tokens
    token 列在 tokenizer 可用时用编码长度填充，否则留 0。
    """
    rows: list[dict] = []
    full_texts, full_idx, cot_texts, cot_idx = [], [], [], []
    try:
        with open(jf, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    obj = json.loads(line)
                except json.JSONDecodeError:
                    continue

                eval_res = obj.get("eval_res")
                eval_details = obj.get("eval_details")
                pred_raw = obj.get("prediction")
                origin_prompt = prediction_val = gold = None
                if isinstance(pred_raw, dict):
                    origin_prompt = pred_raw.get("origin_prompt", "null")
                    prediction_val = pred_raw.get("prediction", "null")
                    gold = pred_raw.get("gold", "null")
                elif isinstance(pred_raw, str):
                    try:
                        parsed = json.loads(pred_raw)
                    except json.JSONDecodeError:
                        parsed = None
                    if isinstance(parsed, dict):
                        origin_prompt = parsed.get("origin_prompt", "null")
                        prediction_val = parsed.get("prediction", "null")
                        gold = parsed.get("gold", "null")
                    else:
                        prediction_val = pred_raw

                i = len(rows)
                rows.append({
                    "eval_res": _fmt_value(eval_res),
                    "eval_details": _fmt_value(eval_details),
                    "origin_prompt": _fmt_value(origin_prompt),
                    "prediction": _fmt_value(prediction_val),
                    "gold": _fmt_value(gold),
                    "full_tokens": 0,
                    "cot_tokens": 0,
                })
                if tokenizer and prediction_val and prediction_val != "null":
                    full_texts.append(prediction_val)
                    full_idx.append(i)
                    cot = _extract_cot(prediction_val)
                    if cot:
                        cot_texts.append(cot)
                        cot_idx.append(i)
    except OSError:
        return None

    if not rows:
        return None

    if tokenizer:
        if full_texts:
            for idx, e in zip(full_idx, tokenizer.encode_batch(full_texts)):
                rows[idx]["full_tokens"] = len(e.ids)
        if cot_texts:
            for idx, e in zip(cot_idx, tokenizer.encode_batch(cot_texts)):
                rows[idx]["cot_tokens"] = len(e.ids)

    wb = Workbook()
    ws = wb.active
    cols = ["eval_res", "eval_details", "origin_prompt", "prediction", "gold", "full_tokens", "cot_tokens"]
    ws.append(cols)
    for r in rows:
        ws.append([r[c] for c in cols])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _add_raw_details(zf: zipfile.ZipFile, rows: list[dict], tokenizer):
    """raw/eval_<id>/ 下只放 *_details.jsonl 转成的 xlsx（含 token 列），不放原始 jsonl。"""
    for r in rows:
        path_str = r.get("details_path")
        if not path_str:
            continue
        base = Path(path_str)
        if not base.exists():
            continue
        for jf in sorted(base.glob("results/**/*_details.jsonl")):
            xlsx = _details_jsonl_to_xlsx(jf, tokenizer)
            if xlsx:
                zf.writestr(f"raw/eval_{r['evaluation_id']}/{jf.stem}.xlsx", xlsx)


def _readme(title: str, view_id: int | None, rows: list[dict]) -> str:
    source = f"分析任务 #{view_id}" if view_id is not None else "临时勾选"
    has_token = any(r.get("token") for r in rows)
    return (
        f"# {title}\n\n"
        f"由测评分析（{source}）导出于 {datetime.utcnow().isoformat()}。\n\n"
        f"包含 {len(rows)} 条 evaluation：\n\n"
        + "\n".join(
            f"- eval#{r['evaluation_id']} · {r['model_name']} · {r['task_key']} "
            f"· {r['version_label']} · accuracy={r['accuracy']}"
            for r in rows
        )
        + "\n\n## 文件说明\n\n"
        "- `summary.xlsx` — 总体对比（任务×模型，分准确率/耗时/Token/CoT）+ 明细两个 sheet\n"
        "- `charts.html` — 静态图表（双击即可在浏览器打开）\n"
        "- `raw/eval_<id>/<task>_details.xlsx` — 逐条样本明细"
        "（eval_res / eval_details / origin_prompt / prediction / gold / full_tokens / cot_tokens）\n"
        + ("" if has_token else "\n> 注：本次未生成 token 统计（tokenizer 不可用或找不到推理产物 jsonl）。\n")
    )
