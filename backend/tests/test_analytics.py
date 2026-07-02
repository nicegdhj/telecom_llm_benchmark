"""AnalysisView CRUD + 导出 zip 测试。

AnalysisView 必须 owner_user_id NOT NULL，所以这里换成"真实用户登录"模式，
而不是 conftest 默认的系统 token。
"""
import io
import json
import zipfile
from datetime import datetime

import pytest
from fastapi.testclient import TestClient

from backend.app.db import get_session
from backend.app.main import app
from backend.app.models import Evaluation, Job, Prediction
from backend.app.services.seed import seed_generic_tasks
from backend.app.services.user_service import create_user


@pytest.fixture
def user_client():
    """以真实用户登录的客户端（id=1，role=admin）。"""
    with get_session() as s:
        from backend.app.models import User
        if not s.query(User).filter_by(username="alice").first():
            create_user(s, "alice", "pw", "admin", None)
        s.commit()
    c = TestClient(app)
    token = c.post("/api/v1/auth/login",
                   json={"username": "alice", "password": "pw"}).json()["session_token"]
    c.headers.update({"Authorization": f"Bearer {token}"})
    return c


def _seed_eval(tmp_path):
    """构造一个完整的 batch + prediction + evaluation；返回 evaluation_id。"""
    from backend.app.models import Batch, BatchCell, Model, Task
    with get_session() as s:
        seed_generic_tasks(s, ["mmlu_redux_gen_5_shot_str"])
        s.commit()
        t = s.query(Task).first()
        m = Model(name="m1", host="h", port=1, model_name="x")
        s.add(m); s.flush()
        b = Batch(name="b1", mode="all")
        s.add(b); s.flush()
        s.add(BatchCell(batch_id=b.id, model_id=m.id, task_id=t.id))
        # 准备 details_path：results/<suite>/xxx_details.jsonl（导出会转成 xlsx）
        details_dir = tmp_path / "details"
        results_dir = details_dir / "results" / "task_43_suite"
        results_dir.mkdir(parents=True)
        with open(results_dir / "task_43_details.jsonl", "w", encoding="utf-8") as f:
            f.write(json.dumps({
                "eval_res": True,
                "eval_details": {"分类结果": 1},
                "prediction": {"origin_prompt": "p", "prediction": "hello", "gold": "g"},
            }, ensure_ascii=False) + "\n")

        jp = Job(type="infer", batch_id=b.id, model_id=m.id, task_id=t.id,
                 params_json={}, status="success", created_at=datetime.utcnow())
        s.add(jp); s.flush()
        p = Prediction(model_id=m.id, task_id=t.id, status="success",
                       num_samples=10, duration_sec=5.0, job_id=jp.id,
                       version_label="v1_infer", finished_at=datetime.utcnow())
        s.add(p); s.flush()
        je = Job(type="eval", batch_id=b.id, model_id=m.id, task_id=t.id,
                 params_json={}, status="success", created_at=datetime.utcnow())
        s.add(je); s.flush()
        e = Evaluation(prediction_id=p.id, eval_version="eval_init", status="success",
                       accuracy=88.5, num_samples=10, duration_sec=0.5,
                       job_id=je.id, version_label="v1_score",
                       details_path=str(details_dir),
                       finished_at=datetime.utcnow())
        s.add(e); s.commit()
        return e.id


def test_create_and_list_analysis_view(user_client, tmp_path):
    eid = _seed_eval(tmp_path)
    r = user_client.post("/api/v1/analysis-views", json={
        "name": "compare_v1",
        "evaluation_ids": [eid],
        "chart_config": {"primary_metric": "accuracy"},
    })
    assert r.status_code == 201
    vid = r.json()["id"]
    assert r.json()["name"] == "compare_v1"

    r2 = user_client.get("/api/v1/analysis-views")
    assert r2.status_code == 200
    assert len(r2.json()) == 1
    assert r2.json()[0]["id"] == vid


def test_get_update_delete(user_client, tmp_path):
    eid = _seed_eval(tmp_path)
    r = user_client.post("/api/v1/analysis-views", json={
        "name": "t", "evaluation_ids": [eid], "chart_config": {}}).json()
    vid = r["id"]

    g = user_client.get(f"/api/v1/analysis-views/{vid}")
    assert g.status_code == 200
    assert g.json()["evaluation_ids"] == [eid]

    u = user_client.put(f"/api/v1/analysis-views/{vid}", json={
        "name": "renamed", "evaluation_ids": [eid], "chart_config": {"k": "v"}})
    assert u.status_code == 200
    assert u.json()["name"] == "renamed"
    assert u.json()["chart_config"] == {"k": "v"}

    d = user_client.delete(f"/api/v1/analysis-views/{vid}")
    assert d.status_code == 204
    assert user_client.get(f"/api/v1/analysis-views/{vid}").status_code == 404


def test_export_returns_zip_with_expected_files(user_client, tmp_path):
    eid = _seed_eval(tmp_path)
    vid = user_client.post("/api/v1/analysis-views", json={
        "name": "export_test", "evaluation_ids": [eid], "chart_config": {}}).json()["id"]
    r = user_client.post(f"/api/v1/analysis-views/{vid}/export")
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/zip"
    assert "export_test.zip" in r.headers.get("content-disposition", "")

    z = zipfile.ZipFile(io.BytesIO(r.content))
    names = z.namelist()
    assert "README.md" in names
    assert "summary.xlsx" in names
    assert "charts.html" in names
    # raw 目录只放 *_details.jsonl 转成的 xlsx
    raw_files = [n for n in names if n.startswith(f"raw/eval_{eid}/")]
    assert any(n.endswith("task_43_details.xlsx") for n in raw_files)
    assert not any(n.endswith(".jsonl") for n in raw_files)
    # 转出的 xlsx 列结构正确
    import openpyxl as _ox
    det = _ox.load_workbook(io.BytesIO(z.read(f"raw/eval_{eid}/task_43_details.xlsx")))
    hdr = list(det.active.iter_rows(values_only=True))[0]
    assert hdr == ("eval_res", "eval_details", "origin_prompt", "prediction", "gold", "full_tokens", "cot_tokens")

    # xlsx 可解析
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(z.read("summary.xlsx")))
    assert "总体对比" in wb.sheetnames
    assert "明细" in wb.sheetnames
    overview = list(wb["总体对比"].iter_rows(values_only=True))
    assert overview[1][0] == "任务"          # 第二行表头第一列
    assert overview[2][1] == 88.5            # 任务行 · 准确率组 · m1 列
    detail = list(wb["明细"].iter_rows(values_only=True))
    assert detail[0][0] == "模型"
    assert detail[1][0] == "m1"              # 模型名
    assert detail[1][5] == 88.5             # 准确率列


def test_export_filename_override(user_client, tmp_path):
    eid = _seed_eval(tmp_path)
    vid = user_client.post("/api/v1/analysis-views", json={
        "name": "default_name", "evaluation_ids": [eid], "chart_config": {}}).json()["id"]
    r = user_client.post(f"/api/v1/analysis-views/{vid}/export?filename=my_report")
    assert r.status_code == 200
    assert "my_report.zip" in r.headers["content-disposition"]


def test_export_strips_path_separators(user_client, tmp_path):
    eid = _seed_eval(tmp_path)
    vid = user_client.post("/api/v1/analysis-views", json={
        "name": "x", "evaluation_ids": [eid], "chart_config": {}}).json()["id"]
    r = user_client.post(f"/api/v1/analysis-views/{vid}/export?filename=../evil/name")
    assert r.status_code == 200
    cd = r.headers["content-disposition"]
    # 取 filename=" ... " 引号里的内容
    import re
    m = re.search(r'filename="([^"]+)"', cd)
    assert m, f"no quoted filename in {cd}"
    assert "/" not in m.group(1)


def test_export_chinese_filename_returns_zip(user_client, tmp_path):
    """Content-Disposition 头部只允许 ASCII；非 ASCII 必须 RFC 5987 编码。"""
    from urllib.parse import quote
    eid = _seed_eval(tmp_path)
    vid = user_client.post("/api/v1/analysis-views", json={
        "name": "x", "evaluation_ids": [eid], "chart_config": {}}).json()["id"]
    r = user_client.post(f"/api/v1/analysis-views/{vid}/export?filename={quote('中文报告')}")
    assert r.status_code == 200
    cd = r.headers["content-disposition"]
    # 有 ASCII fallback + filename* 编码
    assert "filename=" in cd
    assert "filename*=UTF-8''" in cd


def test_export_adhoc_without_saving_template(user_client, tmp_path):
    """临时导出：不先保存模板，直接对一组 evaluation_ids 打包成 zip。"""
    eid = _seed_eval(tmp_path)
    r = user_client.post("/api/v1/analysis-views/export-adhoc",
                         json={"evaluation_ids": [eid], "filename": "临时对比"})
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/zip"
    cd = r.headers.get("content-disposition", "")
    assert "filename*=UTF-8''" in cd  # 中文文件名走 RFC 5987

    z = zipfile.ZipFile(io.BytesIO(r.content))
    names = z.namelist()
    assert "summary.xlsx" in names
    assert "charts.html" in names
    assert any(n.startswith(f"raw/eval_{eid}/") for n in names)


def test_export_adhoc_rejects_empty(user_client):
    r = user_client.post("/api/v1/analysis-views/export-adhoc",
                         json={"evaluation_ids": []})
    assert r.status_code == 400
