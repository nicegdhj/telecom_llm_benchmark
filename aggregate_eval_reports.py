#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
aggregate_eval_reports.py - 汇总 eval_judge.py 产出的评测结果

读取 fmt/ 下所有实验组的 eval_{version}/report.json，
生成与 process_results.py 相同格式的汇总 Excel。

用法：
    python aggregate_eval_reports.py \\
        --fmt-dir /path/to/fmt \\
        --eval-version v3_reeval \\
        --output-dir /path/to/benchmark/outputs

    # 不传 --output-dir 则输出到 fmt-dir/../benchmark/outputs/
"""

import argparse
import concurrent.futures
import glob
import json
import os
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple

import pandas as pd
from openpyxl.styles import PatternFill
from transformers import AutoTokenizer

# ── 模块级辅助函数（供多进程 worker 使用）───────────────────────────────────


def _fmt_value(v):
    """格式化值为字符串，过滤 Excel 非法控制字符。"""
    if v is None:
        return "null"
    if isinstance(v, (dict, list)):
        s = json.dumps(v, ensure_ascii=False, indent=2)
    else:
        s = str(v)
    return re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", s)


# ── 多进程 Worker ──────────────────────────────────────────────────────────

_WORKER_TOKENIZER = None


def _worker_init(tokenizer_path: str | None):
    """在每个 worker 进程中预加载 tokenizer。"""
    global _WORKER_TOKENIZER
    if tokenizer_path and Path(tokenizer_path).exists():
        _WORKER_TOKENIZER = AutoTokenizer.from_pretrained(
            tokenizer_path,
            trust_remote_code=True,
            local_files_only=True,
        )
    else:
        _WORKER_TOKENIZER = None


def _process_suite(args: Tuple) -> Dict[str, Any]:
    """
    处理单个实验组下的一个数据集（suite）。

    参数 (Tuple):
        raw_name, mapped_name, suite, group_dir, eval_version,
        task_mapping, target_dir

    返回:
        {
            "raw_name": str,
            "mapped_suite": str,
            "sample_count": int,
            "file_count": int,
            "token_stats_records": List[Dict],
            "excel_paths": List[str],
            "summary_copies": List[Tuple[str, str]],
        }
    """
    (
        raw_name,
        mapped_name,
        suite,
        group_dir_str,
        eval_version,
        task_mapping,
        target_dir,
    ) = args

    group_dir = Path(group_dir_str)
    eval_results_dir = group_dir / eval_version / "results"
    if not eval_results_dir.exists():
        return {
            "raw_name": raw_name,
            "mapped_suite": get_mapped_suite(suite, task_mapping),
            "sample_count": 0,
            "file_count": 0,
            "token_stats_records": [],
            "excel_paths": [],
            "summary_copies": [],
        }

    suite_dir = eval_results_dir / suite
    if not suite_dir.exists():
        suite_dir = eval_results_dir

    jsonl_files = glob.glob(str(suite_dir / "**" / "*.jsonl"), recursive=True)
    if not jsonl_files:
        jsonl_files = glob.glob(
            str(eval_results_dir / "**" / f"*{suite}*_details.jsonl"),
            recursive=True,
        )
    if not jsonl_files:
        return {
            "raw_name": raw_name,
            "mapped_suite": get_mapped_suite(suite, task_mapping),
            "sample_count": 0,
            "file_count": 0,
            "token_stats_records": [],
            "excel_paths": [],
            "summary_copies": [],
        }

    mapped_suite = get_mapped_suite(suite, task_mapping)
    task_out_dir = os.path.join(target_dir, mapped_suite, raw_name)
    os.makedirs(task_out_dir, exist_ok=True)

    token_stats_records: List[Dict[str, Any]] = []
    excel_paths: List[str] = []
    total_samples = 0

    for jf in jsonl_files:
        rows: List[Dict[str, Any]] = []
        full_texts: List[str] = []
        full_indices: List[int] = []
        cot_texts: List[str] = []
        cot_indices: List[int] = []

        with open(jf, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    obj = json.loads(line)
                except json.JSONDecodeError:
                    continue

                eval_res = obj.get("eval_res")
                eval_details = obj.get("eval_details")
                pred_raw = obj.get("prediction")
                origin_prompt = prediction_val = gold = None

                if isinstance(pred_raw, dict):
                    origin_prompt = pred_raw.get("origin_prompt", "null")
                    prediction_val = pred_raw.get("prediction", "null")
                    gold = pred_raw.get("gold", "null")
                    if origin_prompt in (None, "null"):
                        origin_prompt = _extract_origin_prompt_from_inference_log(
                            pred_raw
                        )
                elif isinstance(pred_raw, str):
                    try:
                        parsed = json.loads(pred_raw)
                        if isinstance(parsed, dict):
                            origin_prompt = parsed.get("origin_prompt", "null")
                            prediction_val = parsed.get("prediction", "null")
                            gold = parsed.get("gold", "null")
                            if origin_prompt in (None, "null"):
                                origin_prompt = _extract_origin_prompt_from_inference_log(
                                    parsed
                                )
                        else:
                            prediction_val = pred_raw
                    except json.JSONDecodeError:
                        prediction_val = pred_raw

                row_idx = len(rows)
                rows.append({
                    "eval_res": _fmt_value(eval_res),
                    "eval_details": _fmt_value(eval_details),
                    "origin_prompt": _fmt_value(origin_prompt),
                    "prediction": _fmt_value(prediction_val),
                    "gold": _fmt_value(gold),
                    "full_tokens": 0,
                    "cot_tokens": 0,
                })

                if _WORKER_TOKENIZER and prediction_val and prediction_val != "null":
                    full_texts.append(prediction_val)
                    full_indices.append(row_idx)
                    cot_text = extract_cot_text(prediction_val)
                    if cot_text:
                        cot_texts.append(cot_text)
                        cot_indices.append(row_idx)

        # ── 批量 token 编码 ───────────────────────────────────────────────
        if _WORKER_TOKENIZER:
            if full_texts:
                encoded = _WORKER_TOKENIZER.backend_tokenizer.encode_batch(full_texts)
                for idx, enc in zip(full_indices, encoded):
                    rows[idx]["full_tokens"] = len(enc.ids)

            if cot_texts:
                encoded = _WORKER_TOKENIZER.backend_tokenizer.encode_batch(cot_texts)
                for idx, enc in zip(cot_indices, encoded):
                    rows[idx]["cot_tokens"] = len(enc.ids)

            for idx in full_indices:
                token_stats_records.append({
                    "experiment": raw_name,
                    "task": mapped_suite,
                    "full_tokens": rows[idx]["full_tokens"],
                    "cot_tokens": rows[idx]["cot_tokens"],
                })

        if rows:
            out_name = os.path.splitext(os.path.basename(jf))[0] + ".xlsx"
            out_path = os.path.join(task_out_dir, out_name)
            pd.DataFrame(rows).to_excel(out_path, index=False)
            excel_paths.append(out_path)
            total_samples += len(rows)

    # 拷贝 summary 子目录
    summary_copies: List[Tuple[str, str]] = []
    suite_summary_src = group_dir / eval_version / "summary" / suite
    if suite_summary_src.exists():
        summary_dst = os.path.join(task_out_dir, "summary")
        if not os.path.exists(summary_dst):
            shutil.copytree(str(suite_summary_src), summary_dst)
            summary_copies.append((str(suite_summary_src), summary_dst))

    return {
        "raw_name": raw_name,
        "mapped_suite": mapped_suite,
        "sample_count": total_samples,
        "file_count": len(jsonl_files),
        "token_stats_records": token_stats_records,
        "excel_paths": excel_paths,
        "summary_copies": summary_copies,
    }


def load_mappings(output_base_dir: str):
    """读取任务名映射和实验组映射（与 process_results.py 相同逻辑）。"""
    task_mapping = {}
    df_task_mapping = None
    task_mapping_file = os.path.join(output_base_dir, "评测任务文件名对应.xlsx")
    if os.path.exists(task_mapping_file):
        df_task_mapping = pd.read_excel(task_mapping_file)
        if len(df_task_mapping.columns) >= 2:
            for _, row in df_task_mapping.iterrows():
                k = str(row.iloc[0]).strip()
                v = str(row.iloc[1]).strip()
                if pd.notna(k) and pd.notna(v) and k != "nan" and v != "nan":
                    task_mapping[k] = v

    id_to_name = {}
    df_exp_mapping = None
    exp_mapping_file = os.path.join(output_base_dir, "实验设置.xlsx")
    if os.path.exists(exp_mapping_file):
        df_exp_mapping = pd.read_excel(exp_mapping_file)
        id_col = "编号" if "编号" in df_exp_mapping.columns else df_exp_mapping.columns[0]
        df_exp_mapping = df_exp_mapping.dropna(subset=[id_col])
        set_col = df_exp_mapping.columns[1]
        for _, row in df_exp_mapping.iterrows():
            try:
                row_id = int(float(row[id_col]))
                set_name = str(row[set_col]).strip()
                if set_name.endswith(".json"):
                    set_name = set_name[:-5]
                id_to_name[row_id] = set_name
            except (ValueError, TypeError):
                continue

    return task_mapping, df_task_mapping, id_to_name, df_exp_mapping


def extract_cot_text(text) -> str:
    """提取 CoT reasoning 文本。

    支持的格式：
      1. <think>...</think>  → 提取标签内内容
      2. ...content...</think> → 只要包含 </think>，前面的全部视为 CoT
      3. 只有 <think> 开头没有结尾（被截断）→ 视为整段都是 CoT
      4. 只有 Thinking Process: 开头没有 </think> 结尾（被截断）→ 视为整段都是 CoT
    """
    if not text:
        return ""
    text = str(text)
    tl = text.lower()

    open_pos = tl.find("<think>")
    close_pos = tl.find("</think>")

    # 1. 完整 <think>...</think>
    if open_pos != -1 and close_pos != -1 and open_pos < close_pos:
        return text[open_pos + 7: close_pos]

    # 2. 包含 </think>，前面全部视为 CoT（兼容 Qwen3.5 等格式）
    if close_pos != -1:
        return text[:close_pos]

    # 3. 只有 <think>，无结尾（被截断）
    if open_pos != -1:
        return text[open_pos + 7:]

    # 4. Thinking Process: 开头（被截断）
    tp_pos = tl.find("thinking process:")
    if tp_pos != -1:
        return text[tp_pos + 17:]

    return ""


def get_mapped_suite(raw_suite: str, task_mapping: dict) -> str:
    """查映射表，找不到时尝试去掉 _suite 后缀再查。"""
    if raw_suite in task_mapping:
        return task_mapping[raw_suite]
    stripped = raw_suite.removesuffix("_suite")
    return task_mapping.get(stripped, raw_suite)


def get_mapped_exp_name(raw_name: str, id_to_name: dict) -> str:
    if raw_name in ("baseline", "baseline_nothink"):
        return raw_name
    if raw_name.startswith("pt"):
        import re
        # 匹配 pt{id}_sft0 或 pt{id}_sf0，并捕获其后的额外后缀（如 _exp0317）
        match = re.match(r"pt(\d+)_sf[t]?0(.*)", raw_name)
        if match:
            pt_id = int(match.group(1))
            extra = match.group(2)  # e.g. "_exp0317" 或 ""
            if pt_id in id_to_name:
                return id_to_name[pt_id] + extra
        # fallback：仅提取 id
        match2 = re.search(r"pt(\d+)_", raw_name)
        if match2:
            pt_id = int(match2.group(1))
            if pt_id in id_to_name:
                return id_to_name[pt_id]
    return raw_name


def _extract_origin_prompt_from_inference_log(pred_dict: dict) -> str:
    """从 inference_log 中提取最后一条 user 消息作为 origin_prompt（兼容 BFCL 等任务）。"""
    try:
        logs = pred_dict.get("inference_log", [])
        if not logs:
            return "null"
        turn = logs[0]
        messages = (
            turn.get("single_turn_inference_data", {}).get("message")
            or turn.get("message")
            or []
        )
        # 取最后一条 role=user 的 content
        for msg in reversed(messages):
            if isinstance(msg, dict) and msg.get("role") == "user":
                content = msg.get("content", "null")
                return content if content else "null"
    except Exception:
        pass
    return "null"


def find_eval_report(group_dir: Path, eval_version: str) -> Path | None:
    """找到 eval_{version}/report.json。"""
    report = group_dir / eval_version / "report.json"
    return report if report.exists() else None


def process(fmt_dir: str, eval_version: str, output_base_dir: str, tokenizer_path: str = None, workers: int = 1):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    fmt_name = Path(fmt_dir).resolve().name  # e.g. "fmt_exp0319"
    target_dir = os.path.join(output_base_dir, f"{fmt_name}_aggregated_reports_{timestamp}")
    os.makedirs(target_dir, exist_ok=True)
    summary_excel_path = os.path.join(target_dir, f"{fmt_name}_总体汇总.xlsx")

    task_mapping, df_task_mapping, id_to_name, df_exp_mapping = load_mappings(output_base_dir)

    tokenizer_exists = tokenizer_path and Path(tokenizer_path).exists()
    if tokenizer_path and not tokenizer_exists:
        print(f"⚠️ tokenizer 路径不存在，跳过 token 统计: {tokenizer_path}")

    # 收集 token 统计（用于后续生成汇总 Sheet）
    token_stats_records = []

    if df_exp_mapping is not None:
        id_col = "编号" if "编号" in df_exp_mapping.columns else df_exp_mapping.columns[0]
        df_exp_mapping = df_exp_mapping.dropna(subset=[id_col])
        new_col = []
        for _, row in df_exp_mapping.iterrows():
            try:
                new_col.append(f"pt{int(float(row[id_col]))}_sft0")
            except Exception:
                new_col.append("ptNaN_sft0")
        df_exp_mapping.insert(0, "实验组编号", new_col)

    # ── 扫描所有实验组 ────────────────────────────────────────────────────────
    all_reports = {}  # (raw_name, mapped_name) -> report data

    for entry in sorted(os.scandir(fmt_dir), key=lambda e: e.name):
        if not entry.is_dir():
            continue
        raw_name = entry.name
        group_dir = Path(entry.path)
        report_path = find_eval_report(group_dir, eval_version)
        if report_path is None:
            print(f"⚠️  跳过 {raw_name}：找不到 {eval_version}/report.json")
            continue

        try:
            data = json.loads(report_path.read_text(encoding="utf-8"))
            mapped_name = get_mapped_exp_name(raw_name, id_to_name)
            all_reports[(raw_name, mapped_name)] = data
        except Exception as e:
            print(f"❌ 读取 {report_path} 失败：{e}")

    if not all_reports:
        print("❌ 未找到任何有效 report.json，请检查 --eval-version 是否正确")
        sys.exit(1)

    print(f"✅ 找到 {len(all_reports)} 个实验组")

    # ── 准备总体对比表 ────────────────────────────────────────────────────────
    # 新 report.json: tasks 是 list，每项有 suite/task/accuracy 字段
    all_suites = set()
    for _, report_data in all_reports.items():
        for t in report_data.get("tasks", []):
            all_suites.add(t.get("suite", t.get("task", "unknown")))

    # baseline 排第一
    exp_groups_sorted = list(all_reports.keys())
    baseline_tuple = next((t for t in exp_groups_sorted if t[0] == "baseline"), None)
    if baseline_tuple:
        exp_groups_sorted.remove(baseline_tuple)
        exp_groups_sorted.insert(0, baseline_tuple)

    task_rows = []
    for raw_suite in sorted(all_suites):
        mapped_suite = get_mapped_suite(raw_suite, task_mapping)
        row_dict = {"Task": mapped_suite}
        for group_tuple in exp_groups_sorted:
            raw_exp = group_tuple[0]
            report_data = all_reports[group_tuple]
            matching = next(
                (t for t in report_data.get("tasks", [])
                 if t.get("suite", t.get("task")) == raw_suite),
                None,
            )
            row_dict[raw_exp] = matching.get("accuracy") if matching else None
        task_rows.append(row_dict)

    df_overview = pd.DataFrame(task_rows)
    cols = ["Task"] + [t[0] for t in exp_groups_sorted]
    df_overview = df_overview[cols]

    # ── 构建并发任务列表 ──────────────────────────────────────────────────────
    suite_tasks = []
    for raw_name, mapped_name in all_reports.keys():
        group_dir = Path(fmt_dir) / raw_name
        report_data = all_reports[(raw_name, mapped_name)]
        suite_set = {t.get("suite", t.get("task")) for t in report_data.get("tasks", [])}
        for suite in suite_set:
            suite_tasks.append((
                raw_name,
                mapped_name,
                suite,
                str(group_dir),
                eval_version,
                task_mapping,
                target_dir,
            ))

    print(f"🚀 开始处理 {len(suite_tasks)} 个数据集（workers={workers}）...")

    # ── 并发处理明细 JSONL ────────────────────────────────────────────────────
    token_stats_records = []
    total_excel = 0
    total_samples = 0
    total_summary = 0

    if workers > 1:
        with concurrent.futures.ProcessPoolExecutor(
            max_workers=workers,
            initializer=_worker_init,
            initargs=(tokenizer_path,),
        ) as executor:
            futures = {
                executor.submit(_process_suite, task): task for task in suite_tasks
            }
            for future in concurrent.futures.as_completed(futures):
                try:
                    result = future.result()
                except Exception as e:
                    task = futures[future]
                    print(f"  ❌ [{task[0]}] {get_mapped_suite(task[2], task_mapping)} 处理失败: {e}")
                    continue

                token_stats_records.extend(result["token_stats_records"])
                total_excel += len(result["excel_paths"])
                total_samples += result["sample_count"]
                total_summary += len(result["summary_copies"])
                if result["sample_count"] > 0:
                    print(
                        f"  ✅ [{result['raw_name']}] {result['mapped_suite']}: "
                        f"{result['file_count']} files, {result['sample_count']} samples"
                    )
    else:
        # 单进程模式（不调 ProcessPoolExecutor，避免序列化开销）
        _worker_init(tokenizer_path)
        for task in suite_tasks:
            result = _process_suite(task)
            token_stats_records.extend(result["token_stats_records"])
            total_excel += len(result["excel_paths"])
            total_samples += result["sample_count"]
            total_summary += len(result["summary_copies"])
            if result["sample_count"] > 0:
                print(
                    f"  ✅ [{result['raw_name']}] {result['mapped_suite']}: "
                    f"{result['file_count']} files, {result['sample_count']} samples"
                )

    print(
        f"\n📊 明细处理完成: {total_excel} 个 Excel, "
        f"{total_samples} 条样本, {total_summary} 个 summary 拷贝"
    )

    # ── 计算 token 聚合统计 ───────────────────────────────────────────────────
    df_token_summary = None
    df_task_token = None
    if tokenizer_exists and token_stats_records:
        print(f"\n📊 计算 token 聚合统计...")
        df_token = pd.DataFrame(token_stats_records)

        # 按 experiment + task 聚合（用于实验组 Sheet 底部）
        df_token_summary = (
            df_token.groupby(["experiment", "task"], dropna=False)
            .agg(
                samples=("full_tokens", "count"),
                full_mean=("full_tokens", "mean"),
                full_median=("full_tokens", "median"),
                full_p90=("full_tokens", lambda x: x.quantile(0.90)),
                full_p95=("full_tokens", lambda x: x.quantile(0.95)),
                full_p99=("full_tokens", lambda x: x.quantile(0.99)),
                full_max=("full_tokens", "max"),
                cot_mean=("cot_tokens", "mean"),
                cot_median=("cot_tokens", "median"),
                cot_p90=("cot_tokens", lambda x: x.quantile(0.90)),
                cot_p95=("cot_tokens", lambda x: x.quantile(0.95)),
                cot_p99=("cot_tokens", lambda x: x.quantile(0.99)),
                cot_max=("cot_tokens", "max"),
            )
            .reset_index()
        )

        int_cols = ["samples", "full_max", "cot_max"]
        for col in int_cols:
            df_token_summary[col] = df_token_summary[col].astype(int)

        float_cols = [
            c for c in df_token_summary.columns
            if c.endswith(("_mean", "_median", "_p90", "_p95", "_p99"))
        ]
        for col in float_cols:
            df_token_summary[col] = df_token_summary[col].round(2)

        # 按 experiment + task 聚合（用于总体对比 Sheet）
        df_task_token = (
            df_token.groupby(["experiment", "task"], dropna=False)
            .agg(full_mean=("full_tokens", "mean"), cot_mean=("cot_tokens", "mean"))
            .reset_index()
        )
        df_task_token["full_mean"] = df_task_token["full_mean"].round(2)
        df_task_token["cot_mean"] = df_task_token["cot_mean"].round(2)

    # 构造总体对比 DataFrame（MultiIndex 列，支持多实验组 + 空列分隔）
    if not df_overview.empty:
        task_names = df_overview["Task"].tolist()
        df_result = pd.DataFrame(index=pd.Index(task_names, name="Task"))
        col_tuples = []

        # 准确率组
        for group_tuple in exp_groups_sorted:
            exp_name = group_tuple[0]
            col_tuples.append(("准确率", exp_name))
            vals = []
            for task in task_names:
                v = df_overview.loc[df_overview["Task"] == task, exp_name].values
                vals.append(v[0] if len(v) > 0 else None)
            df_result[("准确率", exp_name)] = vals

        # 空分隔列
        col_tuples.append(("", ""))
        df_result[("", "")] = None

        # token_full_mean 组
        if df_task_token is not None:
            for group_tuple in exp_groups_sorted:
                exp_name = group_tuple[0]
                col_tuples.append(("token_full_mean", exp_name))
                vals = []
                for task in task_names:
                    v = df_task_token.loc[
                        (df_task_token["experiment"] == exp_name) & (df_task_token["task"] == task),
                        "full_mean"
                    ].values
                    vals.append(v[0] if len(v) > 0 else None)
                df_result[("token_full_mean", exp_name)] = vals
        else:
            for group_tuple in exp_groups_sorted:
                col_tuples.append(("token_full_mean", group_tuple[0]))
                df_result[("token_full_mean", group_tuple[0])] = None

        # 空分隔列
        col_tuples.append(("", ""))
        df_result[("", "")] = None

        # token_COT_mean 组
        if df_task_token is not None:
            for group_tuple in exp_groups_sorted:
                exp_name = group_tuple[0]
                col_tuples.append(("token_COT_mean", exp_name))
                vals = []
                for task in task_names:
                    v = df_task_token.loc[
                        (df_task_token["experiment"] == exp_name) & (df_task_token["task"] == task),
                        "cot_mean"
                    ].values
                    vals.append(v[0] if len(v) > 0 else None)
                df_result[("token_COT_mean", exp_name)] = vals
        else:
            for group_tuple in exp_groups_sorted:
                col_tuples.append(("token_COT_mean", group_tuple[0]))
                df_result[("token_COT_mean", group_tuple[0])] = None

        df_result.columns = pd.MultiIndex.from_tuples(df_result.columns)
        df_overview = df_result

    # ── 生成 Excel ───────────────────────────────────────────────────────────
    with pd.ExcelWriter(summary_excel_path, engine="openpyxl") as writer:
        df_overview.to_excel(writer, sheet_name="总体对比")
        worksheet = writer.sheets["总体对比"]

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        green_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")

        has_baseline = baseline_tuple is not None
        if has_baseline:
            # MultiIndex 表头占 2 行，数据从第 3 行开始
            for row in range(3, len(df_overview) + 3):
                baseline_val = worksheet.cell(row=row, column=2).value
                worksheet.cell(row=row, column=2).fill = yellow_fill
                if isinstance(baseline_val, (int, float)):
                    # 准确率组从第 3 列到 1+len(exp_groups_sorted) 列
                    for col in range(3, 2 + len(exp_groups_sorted)):
                        cell = worksheet.cell(row=row, column=col)
                        if isinstance(cell.value, (int, float)):
                            cell.fill = green_fill if cell.value > baseline_val else red_fill

        current_row = len(df_overview) + 4
        if df_task_mapping is not None:
            worksheet.cell(row=current_row, column=1).value = "【附录 1】任务映射关系"
            df_task_mapping.to_excel(writer, sheet_name="总体对比", startrow=current_row, index=False)
            current_row += len(df_task_mapping) + 3

        # 每个实验组单独一个 Sheet
        for raw_name, mapped_name in all_reports.keys():
            report_data = all_reports[(raw_name, mapped_name)]
            summary = report_data.get("summary", {})
            sum_data = {
                "Type": ["Custom", "Generic"],
                "Count": [
                    summary.get("custom", {}).get("count", 0),
                    summary.get("generic", {}).get("count", 0),
                ],
                "Total Duration (sec)": [
                    summary.get("custom", {}).get("total_duration_sec", 0),
                    summary.get("generic", {}).get("total_duration_sec", 0),
                ],
                "Avg Accuracy": [
                    summary.get("custom", {}).get("avg_accuracy", 0),
                    summary.get("generic", {}).get("avg_accuracy", 0),
                ],
            }
            df_sum = pd.DataFrame(sum_data)

            tasks_list = []
            for t in report_data.get("tasks", []):
                raw_suite = t.get("suite", t.get("task", "unknown"))
                tasks_list.append({
                    "suite": raw_suite,
                    "task": get_mapped_suite(raw_suite, task_mapping),
                    "type": t.get("type", "-"),
                    "eval_type": t.get("eval_type", "-"),
                    "status": t.get("status", "-"),
                    "accuracy": t.get("accuracy"),
                    "num_samples": t.get("num_samples"),
                    "duration_sec": t.get("duration_sec"),
                })
            df_tasks = pd.DataFrame(tasks_list)

            sheet_name = str(raw_name)[:31]
            base = sheet_name
            i = 1
            while sheet_name in writer.sheets:
                sheet_name = f"{base[:28]}_{i}"
                i += 1

            df_sum.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)
            df_tasks.to_excel(writer, sheet_name=sheet_name, index=False, startrow=len(df_sum) + 3)

            # 追加 token 统计到实验组 Sheet 底部
            if df_token_summary is not None:
                exp_token_stats = df_token_summary[df_token_summary["experiment"] == raw_name]
                if not exp_token_stats.empty:
                    ws = writer.sheets[sheet_name]
                    token_start_row = ws.max_row + 2
                    ws.cell(row=token_start_row, column=1, value="【Token 统计】")
                    exp_token_display = exp_token_stats[[
                        "task", "samples", "full_mean", "cot_mean", "full_max", "cot_max"
                    ]]
                    exp_token_display.to_excel(
                        writer, sheet_name=sheet_name, index=False, startrow=token_start_row
                    )

    print(f"📊 总体汇总 Excel: {summary_excel_path}")
    print(f"\n✅ 汇总完成 → {target_dir}")


def main():
    parser = argparse.ArgumentParser(
        description="汇总 eval_judge.py 产出的评测结果为 Excel",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "--fmt-dir",
        default="/Users/jia/MyProjects/pythonProjects/cmcc_cxy/Bprocss/fmt",
        help="fmt 根目录（各实验组的父目录）",
    )
    parser.add_argument(
        "--eval-version",
        required=True,
        help="评测版本号，即 eval_judge.py 的 --eval-version 参数",
    )
    parser.add_argument(
        "--output-dir",
        default=None,
        help="汇总输出目录（默认 fmt-dir/../benchmark/outputs/）",
    )
    parser.add_argument(
        "--tokenizer-path",
        default="./tokenizers/qwen3",
        help="tokenizer 路径，用于统计 prediction 的 token 数",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=1,
        help="并发 worker 数（默认 1，建议设置为 CPU 核心数）",
    )
    args = parser.parse_args()

    output_dir = args.output_dir
    if output_dir is None:
        output_dir = str(Path(args.fmt_dir).parent / "benchmark" / "outputs")

    process(
        fmt_dir=args.fmt_dir,
        eval_version=args.eval_version,
        output_base_dir=output_dir,
        tokenizer_path=args.tokenizer_path,
        workers=args.workers,
    )


if __name__ == "__main__":
    main()
